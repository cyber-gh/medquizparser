import re
import openpyxl

def write_to_excel(question, answers, explanation, correct_answers):
    # Load or create an Excel workbook
    try:
        workbook = openpyxl.load_workbook('mastersheet.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    # Find the last empty row
    last_row = sheet.max_row + 1

    sheet.cell(row=last_row, column=1, value=question)
    for idx, answer in enumerate(answers, 2):
        sheet.cell(row=last_row, column=idx, value=answer)
    sheet.cell(row=last_row, column=9, value=explanation)
    sheet.cell(row=last_row, column=7, value=",".join(str(x) for x in correct_answers))

    workbook.save('mastersheet.xlsx')

def process_correct_answers(question):
    answers = question["answers"]
    correct_answers_idx = []
    for idx in range(len(answers)):
        # print(answers[idx])
        if  answers[idx].strip()[0] == "!":
            correct_answers_idx.append(idx + 1)
    clean_answers = [answer.strip() for answer in answers]
    return {
        "idx": question["idx"],
        "question": question["question"].strip(),
        "answers": clean_answers,
        "correct_answers_idx": correct_answers_idx,
        "explanation": ""
    }


def main():
    f = open("input.txt", "r")
    lines = [line.strip() for line in f.readlines()]
    # pattern = r'(?=[a-f]\))'
    #
    # # Split the lines based on the new pattern
    # lines = [item.strip() for line in lines for item in re.split(pattern, line) if item]

    # lines = [line.strip() for line in lines]

    with open("formatted_input.txt", "w") as fout:
        for line in lines:
            fout.write(line)
            fout.write("\n")

    question_nr = 0
    questions = []
    current_question = {
        "idx": 0,
        "question": "",
        "answers": [],
        "correct_answers_idx": [],
        "explanation": ""
    }

    in_question = False
    for line in lines:

        if len(line) == 0:
            in_question = False
            continue
        if "---" in line:
            in_question = False
            continue
        if "." in line and line.split(".")[0].isnumeric():
            # Start of a new question
            questions.append(process_correct_answers(current_question))
            question_nr = int(line.split(".")[0])
            in_question = True
            current_question = {
                "idx": question_nr,
                "question": " ".join(line.split(".")[1:]),
                "answers": [],
                "correct_answers_idx": [],
                "explanation": ""
            }
        elif line.startswith(("a)", "b)", "c)", "d)", "e)", "f)")):
            ans = line[2:]
            current_question["answers"].append(ans)
        else:
            if 1 <= len(current_question["answers"]) < 4:
                print("Warning -------")
                print(line)
                print("Warning -------")
                current_question["answers"][-1] = current_question["answers"][-1] + "  " + line
            else:
                current_question["question"] = current_question["question"] + " " + line

    questions = questions[1:]
    questions.append(process_correct_answers(current_question))
    for q in questions:
        print(q)
        write_to_excel(q["question"], q["answers"], "", q["correct_answers_idx"])


if __name__ == '__main__':
    main()
