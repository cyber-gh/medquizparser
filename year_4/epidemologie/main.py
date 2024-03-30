import re
import openpyxl

def write_to_excel(question, answers, explanation, correct_answers):
    # Load or create an Excel workbook
    try:
        workbook = openpyxl.load_workbook('mastersheet.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    # Find the last empty row
    last_row = sheet.max_row + 1

    sheet.cell(row=last_row, column=1, value=question)
    for idx, answer in enumerate(answers, 2):
        sheet.cell(row=last_row, column=idx, value=answer)
    sheet.cell(row=last_row, column=9, value=explanation)
    sheet.cell(row=last_row, column=7, value=",".join(str(x) for x in correct_answers))

    workbook.save('mastersheet.xlsx')

def get_cached_answers():
    ans_keys = [0]
    ans_values = [[1]]
    with open("answers.txt", "r") as f:
        lines = [line.strip() for line in f.readlines()]
        for line in lines:
            # if line is a number
            if line[0].isdigit():
                ans_keys.append(int(line.replace(".", "")))
            # else if lines starts with a letter
            elif line[0].isalpha():
                # line is a, b
                # split the line by comma and remove the spaces
                answers = line.split(",")
                answers = [x.strip() for x in answers]
                # map the answers lettes to numbers, a to 1, b to 2, c to 3, etc
                answers = [ord(x) - 96 for x in answers]
                ans_values.append(answers)
            else:
                print("Error: ", line)
    # create a dictionary with the keys and values
    return dict(zip(ans_keys, ans_values))
def process_correct_answers(question):
    answers = question["answers"]
    # sort answers by first letter in the string
    answers.sort(key=lambda x: x[0])

    correct_answers = get_cached_answers()

    # remove first 2 charactesr from each answer
    answers = [answer[2:] for answer in answers]
    return {
        "idx": question["idx"],
        "question": question["question"].strip(),
        "answers": answers,
        "correct_answers_idx": correct_answers[question["idx"]],
        "explanation": ""
    }


def main():
    f = open("input.txt", "r")
    lines = [line.strip() for line in f.readlines()]

    lines = [line for line in lines if line != ""]

    # filter out lines which contain just a number
    lines = [line for line in lines if not line.isdigit()]

    # filter out lines which contain Complement simplu or Complement multiplu
    lines = [line for line in lines if "Complement simplu" not in line]
    lines = [line for line in lines if "Complement multiplu" not in line]

    # filter out lines which start with a number and contain more than one dot like 1.1.2. or 1.2.1.1. or 1.2.
    lines = [line for line in lines if not re.match(r'^\d+\.\d+\.', line)]

    # split out line in multiple lines if it starts with a) b) c) d) e) f)
    # do not split such lines which end with a) or b) or c) or d) or e) or f)

    pattern = r'(?=[a-f]\))'
    lines = [item.strip() for line in lines for item in re.split(pattern, line) if item if line[-2:] not in ["a)", "b)", "c)", "d)", "e)", "f)"]]

    with open("formatted_input.txt", "w") as fout:
        for line in lines:
            fout.write(line)
            fout.write("\n")

    question_nr = 0
    questions = []
    current_question = {
        "idx": 0,
        "question": "",
        "answers": [],
        "correct_answers_idx": [],
        "explanation": ""
    }

    for line in lines:
        # check if lines starts with number with a dot
        if line[0].isdigit() and "." in line:
            # new question starts, store the old one
            questions.append(process_correct_answers(current_question))

            # get all the numbers until the first letter or .
            question_nr = int(re.search(r'\d+', line).group())

            # content of the question is everything after the number
            question = line.removeprefix(str(question_nr)).strip()
            if question.startswith("."):
                question = question[1:].strip()

            current_question = {
                "idx": question_nr,
                "question": question,
                "answers": [],
                "correct_answers_idx": [],
                "explanation": ""
            }
        elif line.startswith(("a)", "b)", "c)", "d)", "e)", "f)")):
            ans = line
            current_question["answers"].append(ans)
        else:
            if 1 <= len(current_question["answers"]) <= 5:
                print("Warning -------")
                print(line)
                print("Warning -------")
                current_question["answers"][-1] = current_question["answers"][-1] + "  " + line
            else:
                current_question["question"] = current_question["question"] + " " + line

    questions = questions[1:]
    questions.append(process_correct_answers(current_question))
    for q in questions:
        print(q)
        write_to_excel(q["question"], q["answers"], "", q["correct_answers_idx"])


if __name__ == '__main__':
    main()
