import re
import openpyxl

def write_to_excel(question, answers, explanation, correct_answers):
    # Load or create an Excel workbook
    try:
        workbook = openpyxl.load_workbook('mastersheet.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    # Find the last empty row
    last_row = sheet.max_row + 1

    sheet.cell(row=last_row, column=1, value=question)
    for idx, answer in enumerate(answers, 2):
        sheet.cell(row=last_row, column=idx, value=answer)
    sheet.cell(row=last_row, column=9, value=explanation)
    sheet.cell(row=last_row, column=7, value=",".join(str(x) for x in correct_answers))

    workbook.save('mastersheet.xlsx')

def process_correct_answers(question):
    answers = question["answers"]
    correct_answers_idx = []
    for idx in range(len(answers)):
        # check if there's an x between [ and ]
        # possible cases cases: [x], [ x], [x ], [ x ]
        if any(x in answers[idx] for x in correct_answers_markers):
            correct_answers_idx.append(idx + 1)
    # remove the markers from the answers
    clean_answers = answers
    for marker in correct_answers_markers:
        clean_answers = [answer.replace(marker, "") for answer in clean_answers]
    clean_answers = [answer.replace("[ ]", "").replace("[x]", "").strip() for answer in clean_answers]
    return {
        "idx": question["idx"],
        "question": question["question"].strip(),
        "answers": clean_answers,
        "correct_answers_idx": correct_answers_idx,
        "explanation": ""
    }


def main():
    f = open("input.txt", "r")
    lines = [line.strip() for line in f.readlines()]

    lines = [line for line in lines if line != ""]

    with open("formatted_input.txt", "w") as fout:
        for line in lines:
            fout.write(line)
            fout.write("\n")

    question_nr = 0
    questions = []
    current_question = {
        "idx": 0,
        "question": "",
        "answers": [],
        "correct_answers_idx": [],
        "explanation": ""
    }

    for line in lines:
        # check if lines starts with number
        if line[0].isdigit():
            # new question starts, store the old one
            questions.append(process_correct_answers(current_question))

            # get all the numbers until the first letter or .
            question_nr = int(re.search(r'\d+', line).group())

            # content of the question is everything after the number
            question = line.removeprefix(str(question_nr)).strip()
            if question.startswith("."):
                question = question[1:].strip()

            current_question = {
                "idx": question_nr,
                "question": question,
                "answers": [],
                "correct_answers_idx": [],
                "explanation": ""
            }
        elif line.startswith(("a)", "b)", "c)", "d)", "e)", "f)")):
            ans = line[2:]
            current_question["answers"].append(ans)
        else:
            if 1 <= len(current_question["answers"]) < 4:
                print("Warning -------")
                print(line)
                print("Warning -------")
                current_question["answers"][-1] = current_question["answers"][-1] + "  " + line
            else:
                current_question["question"] = current_question["question"] + " " + line

    questions = questions[1:]
    questions.append(process_correct_answers(current_question))
    for q in questions:
        print(q)
        write_to_excel(q["question"], q["answers"], "", q["correct_answers_idx"])


if __name__ == '__main__':
    main()
