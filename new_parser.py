import re
from openpyxl import Workbook
from langdetect import detect
import openpyxl

def write_to_excel(question, answers, explanation):
    # Load or create an Excel workbook
    try:
        workbook = openpyxl.load_workbook('mastersheet.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    # Find the last empty row
    last_row = sheet.max_row + 1

    sheet.cell(row=last_row, column=1, value=question)
    for idx, answer in enumerate(answers, 2):
        sheet.cell(row=last_row, column=idx, value=answer)
    sheet.cell(row=last_row, column=9, value=explanation)

    workbook.save('mastersheet.xlsx')


def detect_language(text):
    try:
        return detect(text)
    except:
        return "unknown"


def main():
    f = open("input.txt", "r")
    lines = [line.strip() for line in f.readlines()]

    # lines = [item.strip() for line in lines for item in re.split(r'(?=[A-FА-Ж]\.)', line) if item]
    lines = [item.strip() for line in lines for item in
             (re.split(r'(?=[A-FА-Ж]\.)', line) if not line.startswith(('SC.', 'MC.')) else [line]) if item]

    lines = [item.strip() for line in lines for item in re.split(r'(\d+\.)\s+', line) if item]
    with open("formatted_input.txt", "w") as fout:
        for line in lines:
            fout.write(line)
            fout.write("\n")

    question_nr = 0
    questions = []
    current_question = {
        "idx": 0,
        "question": "",
        "answers": [],
        "explanation": ""
    }
    in_question = False
    for line in lines:
        # If it's a blank line
        if len(line) == 0:
            in_question = False
            pass

        # Checks if it's a nr. row
        if len(line.split(".")) > 0 and line.split(".")[0].isnumeric() and "." in line:
            question_nr = line.split(".")[0]
            if in_question and 0 < len(current_question["answers"]) < 4:
                in_question = True
            else:
                in_question = False
        elif line.startswith("CM.") or line.startswith("CS.") or line.startswith("SC.") or line.startswith(
                "MC.") or line.startswith("СМ.") or line.startswith("CS.") or line.startswith("SC.") or line.startswith(
                "CМ."):
            # now we're parsing a question
            questions.append(current_question)
            in_question = True
            current_question = {"idx": question_nr, "question": line[3:], "answers": [], "explanation": ""}
        elif re.match(r'^[A-FА-Ж]\.', line):
            if in_question:
                current_question["answers"].append(line[2:])
        else:
            if in_question and 0 < len(current_question["answers"]) < 4:
                current_question["answers"][-1] = current_question["answers"][-1] + " " + line
            else:
                if not line.isnumeric():
                    current_question["explanation"] = current_question["explanation"] + line
    f.close()
    idx = 1
    while idx < len(questions) - 2:
        lang = detect_language(questions[idx]["question"])
        if lang == "ro":
            print("RO Question ", questions[idx])
            write_to_excel(questions[idx]["question"], questions[idx]["answers"], questions[idx + 2]["explanation"])
        else:
            print(questions[idx])
        idx += 1



if __name__ == "__main__":
    main()
