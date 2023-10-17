import re
import openpyxl

def write_to_excel(question, answers, explanation, correct_answers):
    # Load or create an Excel workbook
    try:
        workbook = openpyxl.load_workbook('mastersheet.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    # Find the last empty row
    last_row = sheet.max_row + 1

    sheet.cell(row=last_row, column=1, value=question)
    for idx, answer in enumerate(answers, 2):
        sheet.cell(row=last_row, column=idx, value=answer)
    sheet.cell(row=last_row, column=9, value=explanation)
    sheet.cell(row=last_row, column=7, value=",".join(str(x) for x in correct_answers))

    workbook.save('mastersheet.xlsx')

def process_question(question):
    answers = question["answers"]
    clean_answers = []
    correct_answers_idx = []

    in_bold = False
    for idx,answer in enumerate(answers):
        tmp = answer
        if tmp.startswith(("a)", "b)", "c)", "d)", "e)", "f)")):
            tmp = tmp[2:].strip()
        if tmp.startswith("**"):
            tmp = tmp[2:].strip()
            in_bold = True
        if in_bold:
            correct_answers_idx.append(idx + 1)

        if tmp.endswith("**"):
            tmp = tmp[:-2].strip()
            in_bold = False
        if tmp.startswith(("a)", "b)", "c)", "d)", "e)", "f)")):
            tmp = tmp[2:].strip()
        clean_answers.append(tmp)

    return {
        "idx": question["idx"],
        "question": question["question"].strip(),
        "answers": clean_answers,
        "correct_answers_idx": correct_answers_idx,
        "explanation": ""
    }


def is_line_possibly_numeric(line):
    if line.isnumeric():
        return True, int(line)
    if line.endswith("**"):
        line = line[:-2]
    if line.startswith("**"):
        line = line[2:]
    if line.isnumeric():
        return True, int(line)
    else:
        return False, -1

def main():
    f = open("input.txt", "r")
    lines = [line.strip() for line in f.readlines()]
    # pattern = r'(?=[a-f]\))'
    #
    # # Split the lines based on the new pattern
    # lines = [item.strip() for line in lines for item in re.split(pattern, line) if item]

    # lines = [line.strip() for line in lines]

    with open("formatted_input.txt", "w") as fout:
        for line in lines:
            fout.write(line)
            fout.write("\n")

    question_nr = 0
    questions = []
    current_question = {
        "idx": 0,
        "question": "",
        "answers": [],
        "correct_answers_idx": [],
        "explanation": ""
    }

    in_question = False
    for line in lines:

        if len(line) == 0:
            in_question = False
            continue
        if "```" in line:
            in_question = False
            continue
        if "." in line and is_line_possibly_numeric(line.split(".")[0])[0]:
            # Start of a new question
            questions.append(process_question(current_question))
            question_nr = is_line_possibly_numeric(line.split(".")[0])[1]
            in_question = True
            current_question = {
                "idx": question_nr,
                "question": " ".join(line.split(".")[1:]),
                "answers": [],
                "correct_answers_idx": [],
                "explanation": ""
            }
        elif line.startswith(("a)", "b)", "c)", "d)", "e)", "f)")) or line[2:].startswith(("a)", "b)", "c)", "d)", "e)", "f)")):
            ans = line
            current_question["answers"].append(ans)
        else:
            if 1 <= len(current_question["answers"]) < 4:
                print("Warning -------")
                print(line)
                print("Warning -------")
                current_question["answers"][-1] = current_question["answers"][-1] + "  " + line
            else:
                current_question["question"] = current_question["question"] + " " + line

    questions.append(process_question(current_question))
    print(len(questions))
    for q in questions:
        print(q)
        write_to_excel(q["question"], q["answers"], "", q["correct_answers_idx"])


if __name__ == '__main__':
    main()
