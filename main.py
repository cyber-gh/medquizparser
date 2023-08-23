import re
import openpyxl

from test_data import raw_test_data


def parse_text(text):
    current_section_idx = -1
    possible_sections = ["ro", "en", "ru"]
    question = ""
    answers = []
    explanation = ""
    answer_cnt = 0
    for line in text.splitlines():
        line = line.strip()
        if len(line) == 0:
            continue
        if line.split(".")[0].isnumeric():
            continue
        parts = re.split(r'(?=[A-EА-Е]\.)', line)

        for subline in parts:
            print("partial line")
            print(subline)
            if subline.startswith("CM."):
                answer_cnt = 0
                current_section_idx += 1
                if possible_sections[current_section_idx] == "ro":
                    question = subline.split("CM.")[-1].strip()
            elif re.match(r'^[A-FА-Ж]\.', subline):
                answer_cnt += 1
                if possible_sections[current_section_idx] == "ro":
                    answers.append(subline[2:].strip())
            elif answer_cnt >= 4:
                explanation = explanation + subline
            else:
                if possible_sections[current_section_idx] == "ro" and len(answers) != 0:
                    answers[-1] = answers[-1] + " " + subline

    return question, answers, explanation


def write_to_excel(question, answers, explanation):
    # Load or create an Excel workbook
    try:
        workbook = openpyxl.load_workbook('mastersheet.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    # Find the last empty row
    last_row = sheet.max_row + 1

    sheet.cell(row=last_row, column=1, value=question)
    for idx, answer in enumerate(answers, 2):
        sheet.cell(row=last_row, column=idx, value=answer)
    sheet.cell(row=last_row, column=9, value=explanation)

    workbook.save('mastersheet.xlsx')


def test():
    question, answers, explanation = parse_text(raw_test_data)
    print(question)
    print(answers)
    print(explanation)

    write_to_excel(question, answers, explanation)


if __name__ == '__main__':
    while True:
        print("Paste your raw question and press Enter on an empty line to finish:")
        data = []
        while True:
            try:
                line = input()
                if not line:
                    break
                data.append(line)
            except EOFError:
                break

        data = '\n'.join(data)
        if not data:
            print("No data entered. Exiting.")
            break
        question, answers, explanation = parse_text(data)
        print(question)
        print(answers)
        print(explanation)
        write_to_excel(question, answers, explanation)
