import re
import openpyxl

def write_to_excel(question, answers, explanation, correct_answers):
    # Load or create an Excel workbook
    try:
        workbook = openpyxl.load_workbook('mastersheet.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    # Find the last empty row
    last_row = sheet.max_row + 1

    sheet.cell(row=last_row, column=1, value=question)
    for idx, answer in enumerate(answers, 2):
        sheet.cell(row=last_row, column=idx, value=answer)
    sheet.cell(row=last_row, column=9, value=explanation)
    sheet.cell(row=last_row, column=7, value=",".join(str(x) for x in correct_answers))

    workbook.save('mastersheet.xlsx')

def process_correct_answers(question):
    answers = question["answers"]
    correct_answers_idx = []
    for idx in range(len(answers)):
        # print(answers[idx])
        if "[x]" in answers[idx]:
            correct_answers_idx.append(idx + 1)
    clean_answers = [answer.replace("[ ]", "").replace("[x]", "").strip() for answer in answers]
    return {
        "idx": question["idx"],
        "question": question["question"].strip(),
        "answers": clean_answers,
        "correct_answers_idx": correct_answers_idx,
        "explanation": ""
    }


def main():
    f = open("input.txt", "r")
    lines = [line.strip() for line in f.readlines()]
    pattern = r'(?=[a-f]\))'

    # Split the lines based on the new pattern
    lines = [item.strip() for line in lines for item in re.split(pattern, line) if item]

    # Further split lines containing more than 3 dashes in a row at the end
    new_lines = []
    for line in lines:
        dash_pattern = r'(-{3,})$'
        match = re.search(dash_pattern, line)
        if match:
            # Split the line at the position of the dashes
            line_without_dashes = line[:match.start()]
            dashes = match.group(0)
            new_lines.append(line_without_dashes.strip())
            new_lines.append(dashes)
        else:
            new_lines.append(line)

    lines = new_lines

    new_lines = []
    author_pattern = r'(Autor:\s*(Fulga|Globa|Sergiu|sergiu))'
    for line in lines:
        match = re.search(author_pattern, line)
        if match:
            # Split the line after the author
            line_before_author = line[:match.end()]
            line_after_author = line[match.end():]
            new_lines.append(line_before_author.strip())
            if line_after_author:
                new_lines.append(line_after_author.strip())
        else:
            new_lines.append(line)

    lines = new_lines

    lines = [line.strip() for line in lines]

    with open("formatted_input.txt", "w") as fout:
        for line in lines:
            fout.write(line)
            fout.write("\n")

    question_nr = 0
    questions = []
    current_question = {
        "idx": 0,
        "question": "",
        "answers": [],
        "correct_answers_idx": [],
        "explanation": ""
    }

    in_question = False
    for line in lines:

        if len(line) == 0:
            in_question = False
            continue
        if "---" in line:
            in_question = False
            continue
        if "Mod de punctare" in line:
            # Start of a new question
            questions.append(process_correct_answers(current_question))
            question_nr = int(line.split(".")[0])
            in_question = True
            current_question = {
                "idx": question_nr,
                "question": "",
                "answers": [],
                "correct_answers_idx": [],
                "explanation": ""
            }
        elif "Autor:" in line:
            continue
        elif line.startswith(("a)", "b)", "c)", "d)", "e)", "f)")):
            ans = line[2:]
            current_question["answers"].append(ans)
        else:
            if 1 <= len(current_question["answers"]) < 4:
                print("Warning -------")
                print(line)
                print("Warning -------")
                current_question["answers"][-1] = current_question["answers"][-1] + "  " + line
            else:
                current_question["question"] = current_question["question"] + " " + line

    questions = questions[1:]
    questions.append(process_correct_answers(current_question))
    for q in questions:
        print(q)
        write_to_excel(q["question"], q["answers"], "", q["correct_answers_idx"])


if __name__ == '__main__':
    main()
